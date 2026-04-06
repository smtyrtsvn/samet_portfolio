# controls_and_excel.py

from __future__ import annotations
import time
import pyautogui

def press_button(a):

    # Normalize to sequence of keys
    if isinstance(a, str) and '+' in a:
        seq = [k.strip().lower() for k in a.split('+') if k.strip()]
    elif isinstance(a, (list, tuple)):
        seq = [str(k).lower() for k in a]
    else:
        seq = [str(a).lower()]

    # Try pyautogui first
    try:
        import pyautogui
        if len(seq) == 1:
            pyautogui.press(seq[0])
        else:
            pyautogui.hotkey(*seq)
        return True
    except Exception:
        # Fallback to keyboard
        try:
            import keyboard
            combo = '+'.join(seq) if len(seq) > 1 else seq[0]
            keyboard.press_and_release(combo)
            return True
        except Exception as e:
            raise RuntimeError(
                "Key press failed. Install 'pyautogui' or 'keyboard' and ensure a focused window."
            ) from e


def paste(b):
    import pyperclip
    text = "" if b is None else str(b)

    try:
        pyperclip.copy(text)
        print(text, "kopyalandı.")

    except Exception:
        print("copy passed")
        return

    try:
        press_button("ctrl+v")
        print(text, "yapışşştırr")
    except Exception:
        pass


def read_file(path):
    """
    Read an Excel file starting from row 2 (skip the first row) and return
    the content as a list of rows (each row is a list of cell values).

    Supports:
    - .xls  via 'xlrd'
    - .xlsx via 'openpyxl'

    Empty rows are skipped. Empty cells are returned as None.

    Parameters
    ----------
    path : str | pathlib.Path

    Returns
    -------
    list[list]
    """
    from pathlib import Path
    p = Path(path)
    ext = p.suffix.lower()

    if ext == ".xls":
        # Use xlrd (legacy .xls)
        try:
            import xlrd  # xlrd>=2 supports .xls only
        except ImportError as e:
            raise ImportError(
                "Reading .xls requires 'xlrd'. Install with: pip install xlrd"
            ) from e

        book = xlrd.open_workbook(p.as_posix())
        sheet = book.sheet_by_index(0)
        rows = []
        for r in range(1, sheet.nrows):  # start at row index 1 (i.e., Excel row 2)
            # Extract row values
            vals = [sheet.cell_value(r, c) for c in range(sheet.ncols-1)]
            # Normalize empties: xlrd gives '' for empty; convert '' to None
            vals = [None if (v == "" or v is None) else v for v in vals]
            # Skip fully empty rows
            if any(v is not None for v in vals):
                rows.append(vals)
        return rows

    elif ext == ".xlsx":
        # Use openpyxl for modern Excel
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ImportError(
                "Reading .xlsx requires 'openpyxl'. Install with: pip install openpyxl"
            ) from e

        wb = load_workbook(p, data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # start from row 2
            vals = [None if (v == "" or v is None) else v for v in row]
            if any(v is not None for v in vals):
                rows.append(list(vals))
        wb.close()
        return rows

    else:
        raise ValueError("Unsupported file type. Please provide a .xls or .xlsx file path.")


def main():
    print("Starting")
    data = read_file("linesheet.xls")
    print(data)
    print("starting the write time sleep: 5.")

    writer(data)


def writer(data):
    time.sleep(5)
    for row in data:
        for cell in row:
            press_button("ctrl+a")
            time.sleep(0.03)
            paste(str(cell))
            time.sleep(0.03)
            press_button("tab")
            time.sleep(0.03)
main()