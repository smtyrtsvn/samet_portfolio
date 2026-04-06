import openpyxl, glob, os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
orangeFill = PatternFill(start_color='FFAC00', end_color='FFAC00', fill_type='solid')
if __name__ == '__main__':
    filenames = []
    first_run = True
    for file in glob.glob("*.xlsx"):
        filenames.append(file)
    os.chdir("C:\\Users\\Samet Yurtseven\\OneDrive\\Masaüstü\\Oto Koliustu")
    for file in glob.glob("*.xlsx"):  # MAIN LOOP!! File based.
        path1 = "C:\\Users\\Samet Yurtseven\\OneDrive\\Masaüstü\\Oto Koliustu\\" + str(file)
        wb1 = openpyxl.load_workbook(path1, data_only=True)
        sh1 = wb1.active
        cc = 0
        xy = []
        for i in range(5, 200):  # Number of color variations in the packaging list determined.
            color_cell = sh1.cell(row=i, column=5).value
            gram_weight = sh1.cell(row=i, column=6).value
            if color_cell == "Color":
                xy.append(i)  # Get cell row no
            else:
                pass
            if gram_weight == "N.W.":
                weight = sh1.cell(row=i+2, column=6).value # "2159,92"  = net weight of profucts
        order = sh1["E6"].value
        po = order[0:6]
        order = order[7:]
        article = sh1["E7"].value
        for ccc in xy:  #  Loop for colors and variants. For every PL file that have colors/variants inside.
            clr = str("E") + str(ccc+1)
            color = sh1[clr].value
            code = sh1[str("F")+str(ccc+1)].value
            run = True
            r = ccc
            c = 7
            liste = []
            bedenliste = []
            realbeden = []
            cap_list = []
            mixkoli = []
            totalq = 0
            while run:  # Scanned till "Qty". Determine the horizontal borders.
                n = sh1.cell(row=r, column=c).value
                if n == 'Qty/Pack':
                    real_column = c
                    run = False
                else:
                    c = c + 1
            run = True
            while run:  # Scanned till None value. Determine the horizontal borders.
                n = sh1.cell(row=r, column=real_column).value
                if n is None:
                    real_row = r - 1
                    run = False
                else:
                    r = r + 1
            car_qty = sh1.cell(row=real_row, column=3).value
            for row in range(ccc + 1, real_row + 1):            # Ex: From ROW 14 to 32 [14-31].
                for col in range(7, real_column):               # Ex: From COLUMN 7 to 16 [7-15].
                    ik = sh1.cell(row=row, column=1).value      # First Cap.
                    sk = sh1.cell(row=row, column=3).value      # Last Cap.
                    B = sh1.cell(row=ccc, column=col).value     # Size value.
                    A = sh1.cell(row=row, column=col).value     # Quantity of the size in the caps.
                    if B not in bedenliste:
                        bedenliste.append(B)
                    if A is not None:
                        AST = sh1.cell(row=row, column=real_column + 1).value
                        A = int(A) * int(AST)
                        if ik - sk == 0:
                            liste.append([ik, B, A])
                            totalq += A
                            if ik not in cap_list:
                                cap_list.append(ik)
                            else:
                                mixkoli.append(ik)
                        else:
                            for i in range(ik, sk + 1):
                                liste.append([i, B, A])
                                totalq += A
                                if i not in cap_list:
                                    cap_list.append(i)
                                else:
                                    mixkoli.append(i)
            if first_run == True:
                one_gr = float(weight) / totalq
                for i in range(1, sh1.max_row):
                    n = sh1.cell(row=i, column=1).value
                    if n == "Color & Sizes Breakdown:":
                        description = sh1.cell(row=i + 1, column=1).value[len(color) + 3:]
                        first_run = False
                        break
            for i in range(1, 150):  # Cap size info
                n = sh1.cell(row=i, column=7).value
                if n == "L":
                    l = sh1.cell(row=i + 1, column=7).value
                    w = sh1.cell(row=i + 1, column=8).value
                    h = sh1.cell(row=i + 1, column=9).value
                    car_size = str(l) + "x" + str(w) + "x" + str(h)
                    break
            #  Okuma Tamamlandı.
            print(bedenliste, one_gr, "kg / piece.\nQuantity =",totalq)
            if len(bedenliste) < 4:
                path2 = "C:\\Users\\Samet Yurtseven\\Projects\\AutoPyxl\\Koli" + str(4) + "beden.xlsx"
            else:
                path2 = "C:\\Users\\Samet Yurtseven\\Projects\\AutoPyxl\\Koli" + str(len(bedenliste)) + "beden.xlsx"
            wb2 = openpyxl.load_workbook(path2)
            sh2 = wb2.active
            size_col = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            for i in range(0, len(bedenliste)):  # Beden dağılımı ilk sayfaya giriliyor.
                sh2[str(size_col[i]) + '8'].value = bedenliste[i]
                sh2[str(size_col[i]) + '8'].font = Font(bold=True, size=18)
            #  ALL SIZE INFORMATION COMPLETED.
            final_cap = 0
            for i in liste:
                cx = 8 + 30 * (i[0] - 1)
                cx = int(cx)
                try:
                    if i[1] == bedenliste[0]:
                        cy = 'D'
                    elif i[1] == bedenliste[1]:
                        cy = 'E'
                    elif i[1] == bedenliste[2]:
                        cy = 'F'
                    elif i[1] == bedenliste[3]:
                        cy = 'G'
                    elif i[1] == bedenliste[4]:
                        cy = 'H'
                    elif i[1] == bedenliste[5]:
                        cy = 'I'
                    elif i[1] == bedenliste[6]:
                        cy = 'J'
                    elif i[1] == bedenliste[7]:
                        cy = 'K'
                    elif i[1] == bedenliste[8]:
                        cy = 'L'
                    else:
                        print('ELSE ERROR')
                except IndexError:
                    pass
                cxy = str(cy) + str(cx)
                sh2[cxy].value = i[2] # Üstteki yazılıyor.
                sh2[cxy].font = Font(bold=True, size=18)
                fc = size_col[len(bedenliste)]
                if len(bedenliste) <= 4:
                    fc = "H"
                while i[0] in mixkoli:
                    mixkoli.remove(i[0])
                    if len(bedenliste) <= 4:
                        m1 = "F"
                        m2 = ":G"
                        fc = "H"
                    elif len(bedenliste) == 5:
                        m1 = "G"
                        m2 = ":H"
                        fc = "I"
                    elif len(bedenliste) == 6:
                        m1 = "G"
                        m2 = ":I"
                        fc = "J"
                    elif len(bedenliste) == 7:
                        m1 = "H"
                        m2 = ":J"
                        fc = "K"
                    elif len(bedenliste) == 8:
                        m1 = "H"
                        m2 = ":K"
                        fc = "L"
                    else:
                        print("Merge Errorrrrrrrr!!!")
                    mix_cap = m1 + str(cx - 8)  #G2  Üst Merge
                    merged = m1 + str(cx - 8) + m2 + str(cx - 6) #G2:H4
                    sh2.merge_cells(merged)
                    sh2[mix_cap].font = Font(size=48, color="FF0000", bold=True)
                    sh2[mix_cap].fill = orangeFill
                    sh2[mix_cap].value = 'MIX'
                    sh2[mix_cap].alignment = Alignment(horizontal='center', vertical='center')
                    mix_cap = m1 + str(cx + 7)  # Alt Merge
                    merged = m1 + str(cx + 7) + m2 + str(cx + 9)
                    sh2.merge_cells(merged)
                    sh2[mix_cap].font = Font(size=48, color="FF0000", bold=True)
                    sh2[mix_cap].fill = orangeFill
                    sh2[mix_cap].value = 'MIX'
                    sh2[mix_cap].alignment = Alignment(horizontal='center', vertical='center')
                cx += 15
                cxy = str(cy) + str(cx)
                sh2[cxy].value = i[2]  # ALTTAKİNE YAZILIYOR!
                sh2[cxy].font = Font(bold=True, size=18)
                if final_cap <= cx:
                    final_cap = cx
            sh2.delete_rows(cx + 5, 2900)
            sh2.print_area = str("A1:") + str(fc) + str(final_cap + 4)
            info_cell = ["E2", "E3", "E4", "D5", "B6", "D7", "C13", "F14", "A1"]
            info_data = [po, article, '6695', description, order, color, car_size, car_qty, one_gr]
            idk = 0
            for c in info_cell:
                sh2[c].value = info_data[idk]
                sh2[c].font = Font(bold=True, size=18)
                idk += 1
            sh2["M1"].value = one_gr
            sh2["M1"].font = Font(size=10, color="FFFFFF")
    # Dosya aşağıda kaydediliyor.
            error_handler = 0
            if len(color) > 26:
                color = color[:26]
            filename = str(article) + '-' + str(color[6:]) + '-' + str(po) + '-' + str(order)
            os.chdir("C:\\Users\\Samet Yurtseven\\OneDrive\\Masaüstü\\Oto Koliustu\\Koliustu")
            if code is not None:
                filename = str(filename) + "-" + str(code)
            try:
                ffn = "C:\\Users\\Samet Yurtseven\\OneDrive\\Masaüstü\\Oto Koliustu\\Koliustu\\" + str(filename) + '.xlsx'
                wb2.save(ffn)
            except:
                ffn = "C:\\Users\\Samet Yurtseven\\OneDrive\\Masaüstü\\Oto Koliustu\\Koliustu\\" + "Error" + '.xlsx'
                wb2.save(ffn)
            filenames.append(filename)
            print(filename, "saved successfully. \n")
