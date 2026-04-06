import openpyxl, glob, os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def search_all(sh, col_idx, search_string):
    row_space = []
    # In a worksheet, in a specific column, looking for a string and returns row indexes as a list.
    for row in range(1, sh.max_row + 1):
        if sh[row][col_idx].value == search_string:
            row_space.append(row)
        else:
            pass
    return row_space


def main(look4):
    #  Open
    path1 = "C:\\Users\\s.yurtseven\\PycharmProjects\\pythonProject\\AutoPyxl\\OVV.xlsx"
    wb1 = openpyxl.load_workbook(path1, data_only=True)
    sh1 = wb1.active
    # look4 = "FAM0089"  # MODEL NO
    # look4 = input("Çeki Listelerini Oluşturmak İstediğiniz Modelin Kodunu Girin  :")
    Model_index = search_all(sh1, 5, look4)  # Number of Rows that Model shows-up.
    Filenames = []
    err = 1
    All_list = []
    size = ["XS", "S", "M", "L", "XL", "34", "36", "38", "40", "42", "44", "34 / 30", "34 / 32", "36 / 30", "36 / 32", "38 / 30", "38 / 32", "40 / 30", "40 / 32", "42 / 30", "42 / 32", "44 / 30", "Onesize"]
    # size = ["XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL", "86/92", "98/104", "110/116", "122/128", "134/140","146/152", "158/164", "170/176"]
    size_bar = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]
    All_pos = []

    #  Read

    for idx in Model_index:
        model = sh1["F" + str(idx)].value               # All [0]
        color = sh1["I" + str(idx)].value               # All [2]
        c_code = sh1["H" + str(idx)].value              # All [3]
        po_no = sh1["B" + str(idx)].value               # All [1]
        if po_no not in All_pos:
            All_pos.append(po_no)
        portion = sh1["D" + str(idx)].value             # All [4]
        drop = sh1["E" + str(idx)].value                # All [5]
        description = sh1["G" + str(idx)].value         # All [6]
        code = sh1["N" + str(idx)].value                # All [7]
        QUANTITY = []

        for i in range(14, 38):  # 15-38 == "XS-Onesize" 23 values and total at the end as 24th
            q = sh1[idx][i].value
            if q is None:
                q = 0
            QUANTITY.append(q)
        print(QUANTITY)
        All_list.append([model, po_no, color, c_code, portion, drop, description, code, QUANTITY])
                        #  0      1       2       3       4       5       6           7       8
    # print(All_pos)    # >> [767657, 769290, 769254, 769498, 769388, 767582, 767465]

    for p in All_pos:  # POlistesinde, po bazında tek tek tarar. sona kadar burdan devam edebilir. Her sayfa/Po için bir döngü yapar.
        temp_po_list = []
        real_size = []
        for m in All_list:
            if m[1] == p:
                temp_po_list.append(m)
        POqsize = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        #  print(temp_po_list)
        for i in temp_po_list:
            for k in range(0, len(POqsize)):
                POqsize[k] += i[8][k]
        f_size_count = 0
        l_size_count = 0
        for q in POqsize:  # baştan ve sondan kaç basamak 0 tespit edip ona göre beden yerleşimi yapılacak.
            if q == 0:
                f_size_count += 1
            else:
                break
        for q in reversed(POqsize):
            if q == POqsize[-1] or q == 0:
                l_size_count += 1
            else:
                print("l_size_count : ", l_size_count)
                print("f_size_count : ", f_size_count)
                break
        if 24 - (f_size_count + l_size_count) < 4:
            real_size = size[f_size_count: f_size_count+4]
            print("Less than 4 sizes")

        elif 24 - (f_size_count + l_size_count) > 9:
            print("More than 9 sizes!!!")

        else:
            real_size = size[f_size_count: 24 - l_size_count]
            print(real_size)
        #  WRITER
        lenn = len(temp_po_list)
        #  print(temp_po_list)
        if lenn > 4:
            lenn = 1
        path2 = "C:\\Users\\s.yurtseven\\PycharmProjects\\pythonProject\\AutoPyxl\\CheckListTemplate" + str(lenn) + ".xlsx"
        print(path2)
        wb2 = openpyxl.load_workbook(path2)
        sh2 = wb2.active
        sh2["E7"].value = temp_po_list[0][0] #  model
        sh2["E6"].value = str(temp_po_list[0][1]) + " " + str(temp_po_list[0][4]) + " " + str(temp_po_list[0][5])
        col_fill    =    [14, 36, 58, 80]
        qq_fill     =    [41, 64, 87, 110] #  Her seri sonrasında +6 alıyor.
        #  sh2["E9"].value = fabric status percentage
        lenn = len(temp_po_list)
        for n in range(0, lenn): #  Model-Po-Renk-Adet olan bir dizi seçiyor.
            for b in range(0, len(real_size)):
                xy = size_bar[b] + str(col_fill[n]-1)
                sh2[xy].value = real_size[b]
                sh2[xy].font = Font(bold=True, size=10, name='Arial')
            QQ = []
            temp = temp_po_list[n]
            code_color = str(temp[3]) + "-" + str(temp[2])
            sh2[str("E") + str( col_fill[n] )].value = code_color
            sh2[str("A") + str( qq_fill[lenn-1] + 6*(n))].value = code_color + " - " + str(temp[6])
            print(temp[8])
            QQ = temp[8][f_size_count: 24 - l_size_count]
            print("QQ:", QQ)
            if f_size_count + l_size_count == 23:
                print("new code")
                QQ = [temp[8][f_size_count]]
            print(QQ)
            if temp[7] is not None:
                sh2[str("F") + str(col_fill[n])].value = str(temp[7])

            for i in range(0, len(QQ)): # QQ Size sonda Toplam.
                xy = size_bar[i] + str(qq_fill[lenn-1] + 1 + 6*(n))
                sh2[xy].value = QQ[i]
            sh2[str("P") + str(qq_fill[lenn-1]+ 1 + 6*(n))].value = temp[8][-1]

        #  Save
        filename = str(temp_po_list[0][0]) + "-" + str(temp_po_list[0][1]) + "-" + str(temp_po_list[0][4]) + "-" + str(temp_po_list[0][5])
        ffn = "C:\\Users\\s.yurtseven\\Desktop\\Oto Koliustu\\Oto Check List\\" + filename + ".xlsx"
        wb2.save(ffn)


def runner():
    loop_for = ['FAM0563']
# 'FAM0564', 'FAM0565', 'FAM0566', 'FAM0571', 'FAM0572', 'FAM0573', 'FAM0605', 'FAW0652', 'FAW0685', 'FAW0686', 'FAW0691', 'FAW0692'
    for look in loop_for:  # Döngüyü tekrarlamak istediğimiz model kodları burada.
        main(look)
