import mouse
import keyboard
from mouse import MoveEvent, ButtonEvent
from keyboard import KeyboardEvent
import openpyxl
import time
import pyperclip
import pyautogui
from PIL import ImageGrab
from functools import partial
import numpy as np


def detect(img):  # returns the x,y coordinates of the img.
    ImageGrab.grab = partial(ImageGrab.grab, all_screens=True)
    while True:
        screenshot = pyautogui.screenshot()
        screenshot = np.array(screenshot)
        brm = pyautogui.locateOnScreen(img)
        print(brm, type(brm))
        x = int((brm.left + brm.width / 2))
        y = int((brm.top + brm.height / 2))
        print(x, y)
        time.sleep(1)
        if brm is not None:
            break
    return x, y


wb = openpyxl.load_workbook("coretexdata.xlsx", data_only=True)
sh = wb.active
t = time.time() + 0.1

for row in range(2, sh.max_row + 1):
    kod = str(sh.cell(row, 1).value)
    belge_tarih = str(sh.cell(row, 2).value)
    belge_no = str(sh.cell(row, 3).value)
    sevk_tarih = str(sh.cell(row, 4).value)
    tip = str(sh.cell(row, 5).value)
    note = str(sh.cell(row, 6).value)
    atolye_kod = str(sh.cell(row, 7).value)
    order = str(sh.cell(row, 8).value)
    size = [str(sh.cell(row, 9).value),
            str(sh.cell(row, 10).value),
            str(sh.cell(row, 11).value),
            str(sh.cell(row, 12).value),
            str(sh.cell(row, 13).value),
            str(sh.cell(row, 14).value)]
    birim = str(sh.cell(row, 15).value)  # T / E / U / G
    kur = str(sh.cell(row, 16).value)
    fiyat = str(sh.cell(row, 17).value)
    kdv = str(sh.cell(row, 18).value)

    if tip == "g" or tip == "Giden":
        tipx, tipy = 2355, 244
    elif tip == "gg" or tip == "Gelen":
        tipx, tipy = 2355, 264
    else:
        tipx, tipy = 0, 0
        print("tip-x-y-error")

    analiz_events = [[0, 2418, 34], [1],  # Üretim Analiz
                     [0, 2459, 56], [1],  # Üretim Analiz Tablosu
                     [0, 2068, 124], [1],  # + Yeni
                     [0, 2137, 169], [1], [4, tip], [3, "Tab"],  # İşlem tipi
                     [5, belge_tarih], [3, "Tab"],
                     [5, belge_no], [3, "Tab"],
                     [5, atolye_kod], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     [4, "700"], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     [5, order], [3, "Tab"],
                     [0, 2046, 123], [1], ["t", 2],  # Save
                     [0, 2015, 335], [1], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     ]
    for i in size:
        if i != "x":
            analiz_events.append([4, i])
            analiz_events.append([3, "Tab"])

    analiz_events.append([0, 2046, 123])  # saving
    analiz_events.append([1])
    analiz_events.append(["t", 1])

    if tip == "20":  # eğer ticari mal ise, para birimi de girilir.
        analiz_events.append([0, 3273, 335])
        analiz_events.append([1])
        analiz_events.append([2])
        analiz_events.append(["t", 1])
        analiz_events.append([4, birim])
        analiz_events.append(["t", 0.5])
        analiz_events.append([0, 3409, 335])
        analiz_events.append([1])
        analiz_events.append([2])
        analiz_events.append(["t", 0.5])
        analiz_events.append([3, "Bspace"])
        analiz_events.append([5, fiyat])
        analiz_events.append([0, 2046, 123])
        analiz_events.append(["t", 0.5])
        analiz_events.append([1])
        analiz_events.append(["t", 0.5])

    events = analiz_events

    for event in events:  # events loop u. Burada tüm işlemler yapılıyor.
        print(event)
        if event[0] == 0:  # Mouse cursor movement.
            mouse.play([MoveEvent(x=event[1], y=event[2], time=t)])
            t += 0.5
            time.sleep(0.1)
        elif event[0] == 1:  # One click.
            mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.03)])
            t += 0.5
            time.sleep(0.5)
        elif event[0] == 2:  # Dubble click.
            mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.05)])
            mouse.play([ButtonEvent(event_type='down', button='left', time=t + 0.1)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.15)])
            t += 0.5
        elif event[0] == 3:  # Özel Klavye tuşlarının tanıtılması. "TAB" gibi.
            if event[1] == "Tab":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Tab"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Tab")])
                time.sleep(0.5)
            elif event[1] == "g":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="g"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="g")])
                time.sleep(0.1)
            elif event[1] == "Enter":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Enter"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Enter")])
                time.sleep(0.05)
            elif event[1] == "Bspace":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Backspace"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Backspace")])
                time.sleep(0.05)
            elif type(event[1]) == int:  # Sayılar.
                key = str(event[1])
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code=key),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code=key)])
                time.sleep(0.1)
        elif event[0] == 4:  # İstenen kelimenin harflerine tek tek basar.
            temp_data = str(event[1])
            for let in temp_data:  # Kelimenin harfleri tek tek.
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code=str(let)),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code=str(let))])
                time.sleep(0.01)
        elif event[0] == 5:  # Copy - Paste module
            copy_data = str(event[1])
            pyperclip.copy(copy_data)
            keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Control"),  # Press Control
                           KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="V"),  # Press V
                           KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="V"),  # Release V
                           KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Control")])  # Release Control
        elif event[0] == "t":
            time.sleep(int(event[1]))

    print(kod, belge_tarih, belge_no, sevk_tarih, tip, note, atolye_kod, order, size, birim, kur, fiyat, kdv)
