import mouse
import keyboard
from mouse import MoveEvent, ButtonEvent
from keyboard import KeyboardEvent
import openpyxl
import time
import pyperclip

wb = openpyxl.load_workbook("coretexdata.xlsx", data_only=True)
sh = wb.active
t = time.time() + 0.1

for row in range(2, sh.max_row + 1):
    kod = str(sh.cell(row, 1).value)
    belge_tarih = str(sh.cell(row, 2).value)
    belge_no = str(sh.cell(row, 3).value)
    sevk_tarih = str(sh.cell(row, 4).value)
    tip = str(sh.cell(row, 5).value)
    note = str(sh.cell(row, 6).value)
    atolye_kod = str(sh.cell(row, 7).value)
    order = str(sh.cell(row, 8).value)
    size = [str(sh.cell(row, 9).value),
            str(sh.cell(row, 10).value),
            str(sh.cell(row, 11).value),
            str(sh.cell(row, 12).value),
            str(sh.cell(row, 13).value),
            str(sh.cell(row, 14).value),
            str(sh.cell(row, 15).value)]
    birim = str(sh.cell(row, 16).value)  # T / E / U / G
    kur = str(sh.cell(row, 17).value)
    fiyat = str(sh.cell(row, 18).value)
    kdv = str(sh.cell(row, 19).value)
    islem_turu = str(sh.cell(row, 20).value)
    print(islem_turu)

    p_iscilik_events = [[0, 2351, 29], [1],  # Parça İşçilik
                        [0, 2395, 140], [1],  # Parça İşçilik İşlemleri
                        [0, 2068, 124], [1],  # Yeni Parça İşçilik İşlemi Oluştur.
                        # ["t", 1],
                        [0, 2067, 143], [1],  # İşçilik seç
                        [4, kod], [3, "Tab"], [3, "Tab"], [3, "Tab"],  # İşçilik --- 801 - 802 vs.
                        [5, belge_tarih], [3, "Tab"],
                        [5, belge_no], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                        [5, sevk_tarih], [3, "Tab"],
                        [0, 2466, 211], [1], [4, tip], [3, "Enter"], [3, "Tab"],  # İşlem tipi seçimi!
                        [5, note], [3, "Tab"],
                        [5, atolye_kod], [3, "Tab"], [3, "Tab"],
                        [5, order], [3, "Tab"],
                        [0, 2044, 121], [1],
                        #["t", 1]
                        [3, "Tab"], [3, "Enter"],  # Save the report.
                        [0, 2128, 376], [1], [3, "Tab"], [3, "Tab"]  # Select the variant area and go for sizes.
                        ]

    analiz_events = [[0, 2418, 34], [1],  # Üretim Analiz
                     [0, 2459, 56], [1],  # Üretim Analiz Tablosu
                     [0, 3216, 130], [2], # Full Screen
                     [0, 2068, 124], [1],  # + Yeni
                     [0, 2137, 169], [1], [4, tip], [3, "Tab"],  # İşlem tipi
                     [5, belge_tarih], [3, "Tab"],
                     [5, belge_no], [3, "Tab"],
                     [5, atolye_kod], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     [4, "700"], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     [5, order], [3, "Tab"],
                     [0, 2046, 123], [1],  # Save
                     [0, 2015, 335], [1], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     ]

    if islem_turu == "P":  # Parça işçilik yapılıyorsa!
        events = p_iscilik_events  # Coretex rapor türü
        for i in size:
            if i != "x":
                events.append([4, i])
                events.append([3, "Tab"])
        events.append([0, 2046, 123])  # saving
        events.append([1])
        events.append([3, "Enter"])
        events.append(["t", "k"])
        #   events.append([0, 3114, 379])
        #   events.append([2])
        #   events.append([3, "Bspace"])
        #   events.append([5, fiyat])
        #   events.append([3, "Enter"])
        #   events.append([3, "Enter"])
        #   events.append([3, "Enter"])
        events.append([0, 2046, 123])  # saving
        events.append([1])
        events.append([0, 3820, 92])
        events.append([1])
        events.append(["t", 1])

    elif islem_turu == "U":  # Üretim Analiz yapılıyorsa!
        events = analiz_events
        for i in size:
            if i != "x":
                events.append([4, i])
                events.append([3, "Tab"])

        if tip == "20":  # eğer ticari mal ise, para birimi de girilir.
            events.append([0, 3273, 335])
            events.append([1])
            events.append([2])
            events.append(["t", 1])
            events.append([4, birim])
            events.append(["t", 0.5])
            events.append([0, 3409, 335])
            events.append([1])
            events.append([2])
            events.append(["t", 0.5])
            events.append([3, "Bspace"])
            events.append([5, fiyat])
            events.append([0, 2046, 123])
            events.append(["t", 0.5])
            events.append([1])
            events.append(["t", 0.5])

    else:
        print("# # # Error --1--")
        continue



    for event in events:  # events loop u. Burada tüm işlemler yapılıyor.
        print(event)
        if event[0] == 0:  # Mouse cursor movement.
            mouse.play([MoveEvent(x=event[1], y=event[2], time=t)])
            t += 0.5
            time.sleep(0.1)
        elif event[0] == 1:  # One click.
            mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.03)])
            t += 0.5
            time.sleep(0.5)
        elif event[0] == 2:  # Dubble click.
            mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.05)])
            mouse.play([ButtonEvent(event_type='down', button='left', time=t + 0.1)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.15)])
            t += 0.5
        elif event[0] == 3:  # Özel Klavye tuşlarının tanıtılması. "TAB" gibi.
            if event[1] == "Tab":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Tab"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Tab")])
                time.sleep(0.5)
            elif event[1] == "g":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="g"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="g")])
                time.sleep(0.1)
            elif event[1] == "Enter":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Enter"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Enter")])
                time.sleep(0.05)
            elif event[1] == "Bspace":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Backspace"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Backspace")])
                time.sleep(0.05)
            elif type(event[1]) == int:  # Sayılar.
                key = str(event[1])
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code=key),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code=key)])
                time.sleep(0.1)
        elif event[0] == 4:  # İstenen kelimenin harflerine tek tek basar.
            if str(event[1]) != "None" and str(event[1]) is not None:
                temp_data = str(event[1])
                for let in temp_data:  # Kelimenin harfleri tek tek.
                    keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code=str(let)),
                                   KeyboardEvent(event_type=keyboard.KEY_UP, scan_code=str(let))])

            time.sleep(0.01)
        elif event[0] == 5:  # Copy - Paste module
            copy_data = str(event[1])
            pyperclip.copy(copy_data)
            keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Control"),  # Press Control
                           KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="V"),  # Press V
                           KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="V"),  # Release V
                           KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Control")])  # Release Control
        elif event[0] == "t":
            if type(event[1]) is int:
                time.sleep(int(event[1]))
            elif event[1] == "k":
                keyboard.wait("Space")
    print(kod, belge_tarih, belge_no, sevk_tarih, tip, note, atolye_kod, order, size, birim, kur, fiyat, kdv)
