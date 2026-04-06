import mouse
import keyboard
from mouse import MoveEvent, ButtonEvent
from keyboard import KeyboardEvent
import openpyxl
import time
import pyperclip
import pyautogui
from PIL import ImageGrab
from functools import partial
import numpy as np
import win32api


def detect(img):  # returns the x,y coordinates of the img.
    try:
        ImageGrab.grab = partial(ImageGrab.grab, all_screens=True)
        while True:
            Coordinates = pyautogui.locateOnScreen(img)
            print(Coordinates, type(Coordinates))
            x = int((Coordinates.left + Coordinates.width / 2))
            y = int((Coordinates.top + Coordinates.height / 2))
            print(x, y)
            time.sleep(1)
            if type(Coordinates) != "NoneType":
                return x, y
    except:
        pass


wb = openpyxl.load_workbook("coretexdata.xlsx", data_only=True)
sh = wb["Stocker"]
t = time.time() + 0.1
run = False

column_list = []

# iterate through the rows in the selected sheet and append the values in the first column to the list
for row in sh.iter_rows(min_row=1, max_col=1, max_row=sh.max_row):
    for cell in row:
        column_list.append(cell.value)

# print the list
print(column_list)

while run is True:
    a = win32api.GetKeyState(0x01)  # Left mouse button down
    b = win32api.GetKeyState(0x11)  # Left Control Key Pressed
    if a < 0:
        x, y = win32api.GetCursorPos()
        print(x, y)
        time.sleep(0.2)
    if b < 0:
        mouse.play([MoveEvent(x=x, y=y, time=0.2)])
        break
