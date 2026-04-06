import mouse
import keyboard
from mouse import MoveEvent, ButtonEvent
from keyboard import KeyboardEvent
import openpyxl
import time
import pyperclip
import pyautogui
from PIL import ImageGrab
from functools import partial
import numpy as np
import win32api


img_birimfiyat = "C:\\Users\\s.yurtseven\\PycharmProjects\\pythonProject\\PyCoreTex\\imgfolder\\birimfiyat.png"
img_parab = "C:\\Users\\s.yurtseven\\PycharmProjects\\pythonProject\\PyCoreTex\\imgfolder\\parab.png"
img_kur = "C:\\Users\\s.yurtseven\\PycharmProjects\\pythonProject\\PyCoreTex\\imgfolder\\kur.png"
img_durumu = "C:\\Users\\s.yurtseven\\PycharmProjects\\pythonProject\\PyCoreTex\\imgfolder\\durumu.png"

def detect(img):  # returns the x,y coordinates of the img.
    try:
        ImageGrab.grab = partial(ImageGrab.grab, all_screens=True)
        while True:
            Coordinates = pyautogui.locateOnScreen(img)
            print(Coordinates, type(Coordinates))
            x = int((Coordinates.left + Coordinates.width / 2))
            y = int((Coordinates.top + Coordinates.height / 2))
            print(x, y)
            time.sleep(1)
            if type(Coordinates) != "NoneType":
                return x, y
    except:
        pass


wb = openpyxl.load_workbook("coretexdata.xlsx", data_only=True)
sh = wb.active
t = time.time() + 0.1



#  x_pi_1, y_pi_1 = detect("imgfolder\\Piscilik.png")
#  mouse.play([MoveEvent(x=x_pi_1, y=y_pi_1, time=t)])
#  mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
#  mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.03)])
#  t += 0.5

#  x_pi_2, y_pi_2 = detect("imgfolder\\pislem.png")


for row in range(2, sh.max_row + 1):
    start_time = time.time()
    kod = str(sh.cell(row, 1).value)
    belge_tarih = str(sh.cell(row, 2).value)
    belge_no = str(sh.cell(row, 3).value)
    sevk_tarih = str(sh.cell(row, 4).value)
    tip = str(sh.cell(row, 5).value)
    note = str(sh.cell(row, 6).value)
    atolye_kod = str(sh.cell(row, 7).value)
    order = str(sh.cell(row, 8).value)
    size = [str(sh.cell(row, 9).value),
            str(sh.cell(row, 10).value),
            str(sh.cell(row, 11).value),
            str(sh.cell(row, 12).value),
            str(sh.cell(row, 13).value),
            str(sh.cell(row, 14).value),
            str(sh.cell(row, 15).value)]
    birim = str(sh.cell(row, 16).value)  # T / E / U / G
    kur = str(sh.cell(row, 17).value)
    fiyat = str(sh.cell(row, 18).value)
    kdv = str(sh.cell(row, 19).value)
    durumu = str(sh.cell(row, 20).value)  # Sağlam/2.Kalite/Numune -- None/"2"/"n"
    islem_turu = str(sh.cell(row, 21).value)  # U / P

    p_iscilik_events = [[0, 1966, 32], [1],  # Parça İşçilik
                        [0, 2030, 126], [1],  # Parça İşçilik İşlemleri
                        [0, 1659, 111], [1],  # Yeni Parça İşçilik İşlemi Oluştur.
                        ["t", 0.3],
                        [0, 1820, 137], [1],  # İşçilik seç
                        [4, kod], [3, "Tab"], [3, "Tab"], [3, "Tab"],  # İşçilik --- 801 - 802 vs.
                        [5, belge_tarih], [3, "Tab"],
                        [5, belge_no], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                        [5, sevk_tarih], [3, "Tab"],
                        [0, 2091, 186], [1], [4, tip], [3, "Enter"], [3, "Tab"],  # İşlem tipi seçimi!
                        [5, note], [3, "Tab"],
                        [5, atolye_kod], [3, "Tab"], [3, "Tab"],
                        [5, order], [3, "Tab"],
                        [0, 1639, 114], [1], [3, "Tab"], [3, "Enter"],  # Save the report.
                        [0, 1722, 333], [1], [3, "Tab"], [3, "Tab"]  # Select the variant area and go for sizes.
                        ]

    analiz_events = [[0, 2022, 33], [1],  # Üretim Analiz
                     [0, 2021, 58], [1],  # Üretim Analiz Tablosu
                     [0, 2638, 110], [1],  # Full Screen
                     [0, 1653, 114], [1],  # + Yeni
                     [0, 1715, 154], [1], [4, kod], [3, "Tab"],  # İşlem tipi
                     [5, belge_tarih], [3, "Tab"],
                     [5, belge_no], [3, "Tab"],
                     [5, atolye_kod], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     [4, "700"], [3, "Tab"], [3, "Tab"], [3, "Tab"],
                     [5, order], [3, "Tab"],
                     [0, 1640, 110], [1],   # Save
                     [0, 1819, 294], [1], [3, "Tab"]    # Select the first size section.
                     ]


    if islem_turu == "P":  # İslem türü seçimi // Parça İşçilik
        events = p_iscilik_events  # Coretex rapor türü
        for i in size:
            if i != "x":
                events.append([4, i])
                events.append([3, "Tab"])
            else:
                pass
        events.append([0, 1722, 333])  # saving.
        events.append([1])
        events.append([3, "Enter"])
        if durumu == "2":
            events.append(["D", img_durumu, 2, [0, 19]])
            events.append([3, "2"])
            events.append([3, "Tab"])
        elif durumu == "n":
            events.append(["D", img_durumu, 2, [0, 19]])
            events.append([3, "n"])
            events.append([3, "Tab"])
        else:
            pass
        events.append(["D", img_birimfiyat, 2, [0, 19]])
        events.append([3, "Bspace"])
        events.append([5, fiyat])
        events.append([3, "Tab"])
        events.append([3, "Enter"])
        events.append([3, "Enter"])
        events.append([3, "Enter"])
        events.append([0, 1639, 114])  # saving.
        events.append([1])
        events.append([3, "Enter"])
        events.append(["t", 2])  # waits for space key pressed.
        events.append([0, 3050, 85])  # close the window.
        events.append([1])
        events.append(["t", 1])

    elif islem_turu == "U":  # İslem türü seçimi // Üretim Analiz
        events = analiz_events
        for i in size:
            if i != "x":
                events.append([4, i])
                events.append([3, "Tab"])
            else:
                pass
        events.append([0, 1640, 110])  # Save.
        events.append([1])

        if kod == "20":  # Sadece ticari mal alışlar için yapılacak kısım. Aşağısı ve yukarısı aynı. Hata mesajı gelimiyor.
            events.append(["D", img_parab, 2, [0, 19]])
            events.append([4, birim])
            events.append([3, "Tab"])

            events.append(["D", img_kur, 2, [0, 19]])
            events.append([3, "Bspace"])
            events.append([5, kur])
            events.append([3, "Tab"])

            events.append(["D", img_birimfiyat, 2, [0, 19]])  #
            events.append([3, "Bspace"])
            events.append([5, fiyat])
            events.append([3, "Tab"])

        events.append([0, 1640, 110])  # Save.
        events.append([1])
        events.append([0, 3054, 86])  # close the window.
        events.append([1])

    else:
        print("Code Ended --2-- Line 0-175")
        quit()


    print(kod, belge_tarih, belge_no, sevk_tarih, tip, note, atolye_kod, order, size, birim, kur, fiyat, kdv)

    for event in events:  # events loop u. Burada tüm işlemler yapılıyor.
        print(event)
        b = win32api.GetKeyState(0x57)
        if b < 0:
            quit()
        if event[0] == 0:  # Mouse cursor movement.
            mouse.play([MoveEvent(x=event[1]*1.25, y=event[2]*1.25, time=t)])
            t += 0.5
            time.sleep(0.1)
        elif event[0] == 1:  # One click.
            mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.03)])
            t += 0.5
            time.sleep(0.5)
        elif event[0] == 2:  # Dubble click.
            mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.05)])
            mouse.play([ButtonEvent(event_type='down', button='left', time=t + 0.1)])
            mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.15)])
            t += 0.5
        elif event[0] == 3:  # Özel Klavye tuşlarının tanıtılması. "TAB" gibi.
            if event[1] == "Tab":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Tab"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Tab")])
                time.sleep(0.3)
            elif event[1] == "g":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="g"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="g")])
                time.sleep(0.1)
            elif event[1] == "n":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="n"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="n")])
                time.sleep(0.1)
            elif event[1] == "2":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="2"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="2")])
                time.sleep(0.1)
            elif event[1] == "Enter":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Enter"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Enter")])
                time.sleep(0.05)
            elif event[1] == "Bspace":
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Backspace"),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Backspace")])
                time.sleep(0.05)
            elif type(event[1]) == int:  # Sayılar.
                key = str(event[1])
                keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code=key),
                               KeyboardEvent(event_type=keyboard.KEY_UP, scan_code=key)])
                time.sleep(0.1)
        elif event[0] == 4:  # İstenen kelimenin harflerine tek tek basar.
            if str(event[1]) != "None" and str(event[1]) is not None:
                temp_data = str(event[1])
                for let in temp_data:  # Kelimenin harfleri tek tek.
                    keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code=str(let)),
                                   KeyboardEvent(event_type=keyboard.KEY_UP, scan_code=str(let))])

            time.sleep(0.1)
        elif event[0] == 5:  # Copy - Paste module
            copy_data = str(event[1])
            pyperclip.copy(copy_data)
            keyboard.play([KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="Control"),  # Press Control
                           KeyboardEvent(event_type=keyboard.KEY_DOWN, scan_code="V"),  # Press V
                           KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="V"),  # Release V
                           KeyboardEvent(event_type=keyboard.KEY_UP, scan_code="Control")])  # Release Control
            time.sleep(0.1)
        elif event[0] == "t":
            if type(event[1]) is int:
                time.sleep(int(event[1]))
            elif event[1] == "k":
                keyboard.wait("Space")

        elif event[0] == "D":  # Locate the wanted img on screen and get its coordinates. // "D", "image", "click once or twice", "delta [x,y]"//
            img = event[1]
            dx, dy = detect(img)
            dx = dx + event[3][0]
            dy = dy + event[3][1]
            mouse.play([MoveEvent(x=dx, y=dy, time=t)])
            t += 0.5
            time.sleep(0.1)

            if event[2] == 1:  # One click.
                mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
                mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.03)])
                t += 0.5
                time.sleep(0.5)
            elif event[2] == 2:  # Dubble click.
                mouse.play([ButtonEvent(event_type='down', button='left', time=t)])
                mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.05)])
                mouse.play([ButtonEvent(event_type='down', button='left', time=t + 0.1)])
                mouse.play([ButtonEvent(event_type='up', button='left', time=t + 0.15)])
                t += 0.5
    print("--- %s seconds ---" % (time.time() - start_time))